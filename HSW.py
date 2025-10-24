from flask import Flask, render_template_string, request, redirect, url_for
from openpyxl import load_workbook
from datetime import datetime
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
import os, io, json, threading, time

app = Flask(__name__)

# Google Drive setup
FILE_NAME = 'student_roster.xlsx'
LOCAL_EXCEL_PATH = 'student_roster.xlsx'
GOOGLE_CREDENTIALS = json.loads(os.environ['GOOGLE_CREDENTIALS'])
creds = service_account.Credentials.from_service_account_info(GOOGLE_CREDENTIALS)
drive_service = build('drive', 'v3', credentials=creds)

# Globals
file_id = None
file_dirty = False

GROUPS = {
    "Csiga": "🐌",
    "Suni": "🦔",
    "Katica": "🐞",
    "Katica halado": "🐞",
    "Pillango": "🦋",
    "Nyuszi": "🐇",
    "Baglyok": "🦉",
    "Sas": "🦅"
}

def fetch_excel_from_drive():
    """Download the Excel file from Google Drive once at startup."""
    global file_id
    results = drive_service.files().list(q=f"name='{FILE_NAME}'", fields="files(id)").execute()
    if not results['files']:
        raise FileNotFoundError(f"File '{FILE_NAME}' not found in Drive.")
    file_id = results['files'][0]['id']
    request = drive_service.files().get_media(fileId=file_id)
    fh = io.FileIO(LOCAL_EXCEL_PATH, 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    print("✅ Excel file downloaded from Drive")

def upload_excel_to_drive():
    """Upload the updated Excel file back to Drive."""
    global file_dirty
    if not file_dirty:
        return
    media = MediaFileUpload(LOCAL_EXCEL_PATH, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    drive_service.files().update(fileId=file_id, media_body=media).execute()
    file_dirty = False
    print("✅ Excel file uploaded to Drive")

def periodic_sync():
    """Background thread to sync every 15 minutes."""
    while True:
        time.sleep(900)
        upload_excel_to_drive()

threading.Thread(target=periodic_sync, daemon=True).start()

@app.route("/")
def home():
    return render_template_string("""
    <html>
    <head>
      <title>Select Group</title>
      <style>
        body { font-family: sans-serif; background: #f0f4f8; padding: 20px; text-align: center; }
        .group-btn {
          display: inline-block;
          margin: 10px;
          padding: 20px;
          font-size: 24px;
          background: #0078D4;
          color: white;
          border-radius: 10px;
          width: 180px;
          height: 100px;
          line-height: 1.2;
          text-decoration: none;
        }
      </style>
    </head>
    <body>
      <h2>Select a Group</h2>
      {% for name, icon in groups.items() %}
        <a class="group-btn" href="/group/{{ name }}">{{ icon }}<br>{{ name }}</a>
      {% endfor %}
      <p><a href="/sync" style="display:inline-block; margin-top:20px; padding:10px 20px; background:#28a745; color:white; border-radius:5px; text-decoration:none;">🔄 Sync Now</a></p>
    </body>
    </html>
    """, groups=GROUPS)

@app.route("/group/<group>", methods=["GET", "POST"])
def group_view(group):
    global file_dirty
    wb = load_workbook(LOCAL_EXCEL_PATH)
    if group not in wb.sheetnames:
        return f"<h3>Group '{group}' not found.</h3>", 404

    ws = wb[group]
    today_str = datetime.now().strftime("%d/%m/%Y")

    date_col = None
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=1, column=col).value
        if isinstance(cell, datetime):
            cell = cell.strftime("%d/%m/%Y")
        if str(cell) == today_str:
            date_col = col
            break
    if date_col is None:
        return f"<h3>No check-in column found for today ({today_str}).</h3>", 400

    students = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name not in [None, ""]:
            checked_in = False
            if date_col:
                cell = ws.cell(row=row, column=date_col)
                checked_in = (cell.value == "✅")
            students.append({"name": name, "checked_in": checked_in})

    if request.method == "POST":
        student = request.form["student"]
        if date_col:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == student:
                    cell = ws.cell(row=row, column=date_col)
                    if cell.value != "✅":
                        cell.value = "✅"
                        wb.save(LOCAL_EXCEL_PATH)
                        file_dirty = True
                    break
        return redirect(url_for("group_view", group=group, checked=student))

    checked = request.args.get("checked", "")
    return render_template_string("""
    <html>
    <head>
      <title>{{ group }} Students</title>
      <style>
        body { font-family: sans-serif; background: #f0f4f8; padding: 20px; text-align: center; }
        .student-grid { display: flex; flex-wrap: wrap; justify-content: center; gap: 15px; }
        .student-card {
          width: 120px;
          height: 140px;
          border-radius: 10px;
          box-shadow: 0 0 5px rgba(0,0,0,0.1);
          background: white;
          padding: 0;
          opacity: 1;
        }
        .student-card.checked {
          opacity: 0.4;
        }
        .card-button {
          width: 100%;
          height: 100%;
          border: none;
          background: none;
          font-size: 24px;
          padding: 10px;
          cursor: pointer;
          display: flex;
          flex-direction: column;
          justify-content: center;
          align-items: center;
          color: #333;
        }
        .card-button:hover {
          background: #e0f0ff;
          border-radius: 10px;
        }
        .card-button p {
          margin: 5px 0 0;
          font-size: 16px;
        }
      </style>
    </head>
    <body>
      <h2>{{ group }} {{ icon }}</h2>
      {% if checked %}
        <p style="color: green;">{{ checked }} checked in successfully!</p>
      {% endif %}
      <div class="student-grid">
        {% for student in students %}
          <form method="POST" class="student-card {% if student.checked_in %}checked{% endif %}">
            <input type="hidden" name="student" value="{{ student.name }}">
            <button type="submit" class="card-button" {% if student.checked_in %}disabled{% endif %}>
              <div style="font-size: 40px;">👤</div>
              <p>{{ student.name }}</p>
            </button>
          </form>
        {% endfor %}
      </div>
      <p><a href="/">← Back to Groups</a></p>
    </body>
    </html>
    """, group=group, icon=GROUPS.get(group, ""), students=students, checked=checked)

@app.route("/sync")
def sync_now():
    try:
        upload_excel_to_drive()
        return "<h3>✅ Synced to Google Drive successfully.</h3><p><a href='/'>← Back to Home</a></p>"
    except Exception as e:
        return f"<h3>❌ Sync failed: {e}</h3><p><a href='/'>← Back to Home</a></p>"

if __name__ == "__main__":
    fetch_excel_from_drive()
    app.run(host="0.0.0.0", port=5000)
